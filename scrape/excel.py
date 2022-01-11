from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

wb = Workbook()
ws = wb.active
ws.title = "CVE"


def zrob_tabelke():
    for host in hosts:
        zmienna = ""
        #print(host.get('id'))
        host_name = host.get('id')
        if host_name == 'container' or host_name == 'menubox':
            continue
        else:
            host_name = host_name.split('_')[1]
            #print(host_name)
            #cve.write(host_name + '\n')
        
        scripts = host.find_all('tr', attrs={'class': 'script'})
        #print(script)
        #toco = script.find('td', colspan='6')
        for sc in scripts:
            table_data = sc.find_all('td')
            vuln = False
            for td1 in table_data:
                if td1.text.startswith('vulners'):
                    print(td1.text)
                    vuln = True
            if vuln == True:
                toco = sc.find('pre')
                t1 = toco.text.splitlines()
                for t2 in t1:
                    if t2.lstrip().startswith('CVE'):
                        if float(t2.lstrip().split()[1]) > 5.0:
                            print(t2.lstrip())
                            ws.append(t2.lstrip().split())
                            #zmienna = zmienna + t2.lstrip() + '\n'
                            #cve.write(t2.lstrip()+'\n')
                #print(t1.lstrip())
                #cve.write('###############'+ '\n')         
                zmienna = zmienna + '###############'+ '\n'         
                
                #zmienna = zmienna + '########################################'+ '\n'
            #cve.write('########################################'+ '\n')
            #cve.write('########################################'+ '\n')
        if zmienna:
            zmienna = zmienna + '########################################'+ '\n'
            zmienna = zmienna + '\n' + '___'+ '\n'
            zmienna ='# '+ host_name + '\n' + zmienna
            zmienna = zmienna.splitlines()
            print(zmienna)
            ws.insert_rows(1)
            ws['A1'].value = host_name
            #ws.append(zmienna)
            #cve.write(zmienna)    
    wb.save('cve.xlsx')
    """
            toco = sc.find('pre')
            t1 = toco.text.splitlines()
            for t2 in t1:
                if t2.lstrip().startswith('CVE'):
                    if float(t2.lstrip().split()[1]) > 5.0:
                        print(t2.lstrip())
                        cve.write(t2.lstrip()+'\n')
            #print(t1.lstrip())
            cve.write('###############'+ '\n')
        cve.write('########################################'+ '\n')    
        cve.write('########################################'+ '\n')    
        
    """

def wypisz_CVE():
    for host in hosts:
        zmienna = ""
        #print(host.get('id'))
        host_name = host.get('id')
        if host_name == 'container' or host_name == 'menubox':
            continue
        else:
            host_name = host_name.split('_')[1]
            #print(host_name)
            #cve.write(host_name + '\n')
        
        scripts = host.find_all('tr', attrs={'class': 'script'})
        #print(script)
        #toco = script.find('td', colspan='6')
        for sc in scripts:
            table_data = sc.find_all('td')
            vuln = False
            for td1 in table_data:
                if td1.text.startswith('vulners'):
                    print(td1.text)
                    vuln = True
            if vuln == True:
                toco = sc.find('pre')
                t1 = toco.text.splitlines()
                for t2 in t1:
                    if t2.lstrip().startswith('CVE'):
                        if float(t2.lstrip().split()[1]) > 5.0:
                            print(t2.lstrip())
                            zmienna = zmienna + t2.lstrip() + '\n'
                            #cve.write(t2.lstrip()+'\n')
                #print(t1.lstrip())
                #cve.write('###############'+ '\n')         
                zmienna = zmienna + '###############'+ '\n'         
                
                #zmienna = zmienna + '########################################'+ '\n'
            #cve.write('########################################'+ '\n')
            #cve.write('########################################'+ '\n')
        if zmienna:
            zmienna = zmienna + '########################################'+ '\n'
            zmienna = zmienna + '\n' + '___'+ '\n'
            zmienna ='# '+ host_name + '\n' + zmienna
            #cve.write(zmienna)    

        """
            toco = sc.find('pre')
            t1 = toco.text.splitlines()
            for t2 in t1:
                if t2.lstrip().startswith('CVE'):
                    if float(t2.lstrip().split()[1]) > 5.0:
                        print(t2.lstrip())
                        cve.write(t2.lstrip()+'\n')
            #print(t1.lstrip())
            cve.write('###############'+ '\n')
        cve.write('########################################'+ '\n')    
        cve.write('########################################'+ '\n')    
        
        """


with open('vulners.html','r') as html_file:
    content = html_file.read()

    soup = BeautifulSoup(content, 'lxml')
    hosts = soup.find_all('div')
    zrob_tabelke()
#    print(hosts.get('id'))
    #wypisz_CVE()